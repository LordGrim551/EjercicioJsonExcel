import json
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

fecha_actual = datetime.now()

# Obtener el mes y el año de la fecha actual
mes_actual = fecha_actual.strftime("%B")
año_actual = fecha_actual.strftime("%Y")
nombre_archivo_excel = f"pagos-empleados-{mes_actual.lower()}-{año_actual}.xlsx"

with open('employees.json', 'r') as f:
    data = json.load(f)

Salario = [float(Salary['salary'].replace('$', '').replace(',', '')) for Salary in data]
Edad = [Age['age'] for Age in data]
Nombre = [Name['name'] for Name in data]
Genero = [Gender['gender'] for Gender in data]
Proyecto = [Proyect['proyect'] for Proyect in data]
Email = [Email['email'] for Email in data]

df = pd.DataFrame({'Salario': Salario, 'Edad': Edad,
                   'Nombre': Nombre, 'Genero': Genero,
                   'Proyecto': Proyecto, 'Email': Email})

for index, row in df.iterrows():
    if row['Edad'] < 30:
        df.at[index, 'Salario'] *= 1.10

for index, row in df.iterrows():
    if row['Proyecto'] == 'GRONK':
        """inplace se aplicaran los cambios de no mostrar"""
        df.drop(index, inplace=True)
# Convertir el símbolo del salario a euro (€)
# Se utiliza una función lambda para agregar el símbolo de euro (€) al final de cada valor en la columna 'Salario'.
df['Salario'] = df['Salario'].map(lambda x: f"{x}€")

print(df)

df.index.name = 'ID'
df.to_excel(nombre_archivo_excel)

wb = load_workbook(nombre_archivo_excel)
ws = wb.active
for row in ws.iter_rows():
    for cell in row:
        print(cell.value)
